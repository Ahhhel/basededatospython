from tkinter import *    # Carga módulo tk (widgets estándar)
from tkinter import ttk,font  # Carga ttk (para widgets nuevos 8.5+)}
from openpyxl import *
#se importa la libreria openpyxl para lectura de archivos excel >= 2010 

class Aplicacion():
    #opciones de la interfaz
    def __init__(self):
       self.raiz = Tk()
       fon = font.Font(size=16) #se crea una fuente para la letra
       s = ttk.Style()
       s.configure('my.TButton', font=('Helvetica', 14))
       self.raiz.title("Lector de archivos") #titulo de la ventana
       self.raiz.geometry("1280x600") #tamaño de la ventana

       self.areatext = ttk.Label(self.raiz, text= "Información", font=fon) #etiqueta de info
       self.tinfo = Text(self.raiz, width=80, height=20, font=fon)#espacio donde irá la informacion obtenida del archivo
       self.boton2 = ttk.Button(self.raiz, text = "Ejecutar consulta", style='my.TButton', command = self.lectura_archivo)
       self.boton1 = ttk.Button(self.raiz, text="Salir", style='my.TButton', command = quit) #boton salir

       self.areatext.pack(side=TOP)
       self.tinfo.pack(side=TOP)  
       self.boton2.pack(side=TOP, fill=BOTH)
       self.boton1.pack(side=BOTTOM, fill=BOTH)
       
       self.raiz.mainloop() #funcion del loop de la interfaz
    def consulta_tabla(self):
        wb = load_workbook('Algebra relacional.xlsx') #se hace la apertura del archivo
        ws = wb.active #se activa el archivo
        for row in ws.rows:
            for col in row:

                print(col.value),
        print("")
    #aqui puede ir el back en otra funcion
    def lectura_archivo(self):
        texto = "\n"
        self.tinfo.delete("1.0", END)
        wb = load_workbook('Algebra relacional.xlsx') #se hace la apertura del archivo
        ws = wb.active #se activa el archivo
        z = 0
        col1 = 0
        col2 = 0
        for row in ws.iter_rows(): #este for solo cuenta cuantas filas son
            z += 1
        for col in ws.iter_cols(): #este for es inutil realmente, pero ps idenfica las columnas de NAME y salary
            col1 += 1
            col2 += 1
            for cell in col:
                if cell.value == "FIRST_NAME":
                    cmpo = cell.value
                    fincol1 = col1
                if cell.value == "SALARY":
                    cmpo2 = cell.value
                    fincol2 = col2
        # print(cmpo, cmpo2)
        texto += str(cmpo) + "\t \t" + str(cmpo2) + "\n"
        for x in range (2,z): #este for es el que está chido ya que solo recorremos las filas a partir de la 2 
                                #ya que la fila 1 solo es de entidades
            if (int(ws.cell(row=x,column=fincol2).value) >= 10000 and int(ws.cell(row=x,column=fincol2).value) <= 15000) :
                texto += str(ws.cell(row=x,column=fincol1).value) + "\t \t" + str(ws.cell(row=x,column=fincol2).value) + "\n"
        self.tinfo.insert("1.0", texto)
        
""" mi_app = Aplicacion()
mi_app.lectura_archivo() """

def main():
    mi_app= Aplicacion()
    return 0

if __name__ == '__main__':
    main()   